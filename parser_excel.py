import time
from random import choice, randint
from typing import Iterable, Mapping, Any, Generator


from dotenv import dotenv_values
from openpyxl import load_workbook
from psycopg2 import connect


def to_postgresql_database(placements: Iterable[Mapping[str, Any]],
                           table_name: str,
                           host: str, user: str,
                           password: str,
                           database: str, port: str) -> None:
    """
    Создаёт из итерируемого объекта и имён столбцов
    sql таблицу в существующей базе данных."
    """
    table_name = table_name.lower().replace(' ', '_')

    with connect(host=host, user=user,
                 password=password,
                 database=database, port=port) as connection:
        connection.autocommit = True

        with connection.cursor() as cursor:
            cursor.execute(
                f"DROP TABLE IF EXISTS {table_name};"
            )

            cursor.execute(
                f"""CREATE TABLE {table_name}
                (id                              INT GENERATED BY DEFAULT AS IDENTITY PRIMARY KEY,
                placement                        VARCHAR(30)  NOT NULL,
                number_of_the_room               INT          NOT NULL,
                type_of_placement                VARCHAR(20)  NOT NULL,
                status                           VARCHAR(30)  NOT NULL,
                new_building_or_resale           VARCHAR(20)  NOT NULL,
                placement_area                   FLOAT(2)     NOT NULL,
                floor                            INT          NOT NULL,
                living_area                      FLOAT(2)     NOT NULL,
                number_of_rooms                  VARCHAR(30)  NOT NULL,
                total_area                       FLOAT(2)     NOT NULL,
                area_without_balconies_loggias   FLOAT(2)     NOT NULL,
                cadastral_number                 VARCHAR(30)  NOT NULL,
                created_at                       TIMESTAMP    NOT NULL,
                updated_at                       TIMESTAMP    NOT NULL);"""
            )

            date_and_time = time.time()
            for placement in placements:
                cursor.execute(
                    f"""INSERT INTO placements VALUES 
                    (DEFAULT,
                    '{placement["Помещение"]}',
                    '{placement["Номер помещения"]}',
                    '{placement["Вид помещения"]}',
                    '{placement["Статус"]}',
                    '{placement["Новостройка/Вторичка"]}',
                    {placement["Площадь помещения"]},
                    {placement["Этаж"]},
                    {placement["Жилая площадь"]},
                    '{placement["Количество комнат"]}',
                    {placement["Общая площадь"]},
                    {placement["Площадь без балконов/лоджий"]},
                    '{placement["Кадастровый номер"]}',
                    to_timestamp('{date_and_time}'),
                    to_timestamp('{date_and_time}'))"""
                )


def get_data() -> Generator:
    wb = load_workbook('таблица с данными.xlsx')
    sheet = wb.active

    row = 10
    while sheet.cell(row=row, column=1).value:
        placement = {'Помещение':
                         sheet.cell(row=row, column=1).value,
                     'Номер помещения':
                         int(sheet.cell(row=row, column=1).value.split('№')[1]),
                     'Вид помещения':
                         "Нежилое" if sheet.cell(row=row, column=1).value.split()[0] == "Нежилое"
                         else "Жилое",
                     'Статус':
                         sheet.cell(row=row, column=4).value,
                     'Новостройка/Вторичка':
                         choice(["Новостройка", "Вторичка"]),
                     'Площадь помещения':
                         float(sheet.cell(row=row, column=5).value.replace(',', '.'))
                         if sheet.cell(row=row, column=5).value else 0,
                     'Этаж':
                         int(sheet.cell(row=row, column=7).value),
                     'Жилая площадь':
                         float(sheet.cell(row=row, column=9).value.replace(',', '.'))
                         if sheet.cell(row=row, column=9).value else 0,
                     'Количество комнат':
                         sheet.cell(row=row, column=10).value,
                     'Общая площадь':
                         float(sheet.cell(row=row, column=11).value.replace(',', '.')),
                     'Площадь без балконов/лоджий':
                         float(sheet.cell(row=row, column=12).value.replace(',', '.'))
                         if sheet.cell(row=row, column=12).value else 0,
                     'Кадастровый номер':
                         f"02:57:020102:{randint(0, 10000)}"
                     }
        yield placement
        row += 1


def main():
    data = get_data()
    config = dotenv_values(".env")
    to_postgresql_database(
        data, "placements",
        host=config['HOST'],
        user=config['POSTGRES_USER'],
        password=config['POSTGRES_PASSWORD'],
        database=config['POSTGRES_DB'],
        port=config['PORT'],
    )


if __name__ == '__main__':
    main()
